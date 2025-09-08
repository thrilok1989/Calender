import os
import math
import json
import time
import pathlib
import streamlit as st
import threading
from datetime import datetime, timedelta
from typing import Optional, Tuple, Dict

import numpy as np
import pandas as pd
import requests

# ===================== CONFIG =====================
SYMBOL = os.getenv("NSE_SYMBOL", "NIFTY").upper()  # "NIFTY" or "BANKNIFTY"
EXCEL_FILE = os.getenv("EXCEL_FILE", f"expiry_day_{SYMBOL.lower()}_analysis.xlsx")
SNAPSHOT_DIR = os.getenv("SNAPSHOT_DIR", f"snapshots_{SYMBOL.lower()}")
pathlib.Path(SNAPSHOT_DIR).mkdir(parents=True, exist_ok=True)

ZONE_WIDTH = int(os.getenv("ZONE_WIDTH", 10))
GAMMA_BREACH = int(os.getenv("GAMMA_BREACH", 30))
FAR_FROM_MAXPAIN = int(os.getenv("FAR_FROM_MAXPAIN", 100))
PCR_BULL = float(os.getenv("PCR_BULL", 1.2))
PCR_BEAR = float(os.getenv("PCR_BEAR", 0.9))
TOP_N_DELTA_OI = int(os.getenv("TOP_N_DELTA_OI", 5))
HTTP_TIMEOUT = int(os.getenv("HTTP_TIMEOUT", 20))

# Telegram Config (optional)
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID", "")

BASE = "https://www.nseindia.com"
OC_URL = f"{BASE}/api/option-chain-indices?symbol={SYMBOL}"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/123.0.0.0 Safari/537.36"
    ),
    "accept-language": "en-US,en;q=0.9",
    "accept-encoding": "gzip, deflate, br",
}

# ===================== TELEGRAM FUNCTIONS =====================
def send_telegram_message(message):
    """Send message via Telegram bot (optional feature)"""
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        return False
    
    try:
        # Try to import telegram, but don't fail if not available
        try:
            import telegram
        except ImportError:
            st.warning("Telegram package not installed. Install with: pip install python-telegram-bot")
            return False
            
        bot = telegram.Bot(token=TELEGRAM_BOT_TOKEN)
        bot.send_message(chat_id=TELEGRAM_CHAT_ID, text=message)
        return True
    except Exception as e:
        st.error(f"Failed to send Telegram message: {e}")
        return False

# ===================== TIME VALIDATION =====================
def is_valid_trading_time():
    """Check if current time is within trading hours (Mon-Fri, 9:00-15:30)"""
    now = datetime.now()
    
    # Check if weekday (Monday=0, Friday=4)
    if now.weekday() > 4:
        return False
    
    # Check time (9:00 to 15:30)
    current_time = now.time()
    start_time = datetime.strptime("09:00", "%H:%M").time()
    end_time = datetime.strptime("15:30", "%H:%M").time()
    
    return start_time <= current_time <= end_time

# ===================== DATA FETCH =====================
def _nse_session() -> requests.Session:
    s = requests.Session()
    # Seed cookies: visit homepage and option-chain page once
    try:
        s.get(BASE, headers=HEADERS, timeout=HTTP_TIMEOUT)
    except:
        pass
    return s

def fetch_option_chain() -> dict:
    s = _nse_session()
    r = s.get(OC_URL, headers=HEADERS, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    return r.json()

# ===================== PARSE & CORE METRICS =====================
def parse_chain(raw: dict) -> Tuple[pd.DataFrame, float, str, str]:
    records = raw.get("records", {})
    filtered = raw.get("filtered", {})

    spot = float(records.get("underlyingValue") or filtered.get("underlyingValue") or 0.0)
    expiries = records.get("expiryDates") or filtered.get("expiryDates") or []
    expiry = expiries[0] if expiries else ""
    rows = []
    for item in records.get("data", []):
        sp = item.get("strikePrice")
        if sp is None:
            continue
        ce = item.get("CE", {})
        pe = item.get("PE", {})
        rows.append({
            "strikePrice": float(sp),
            "CE_OI": int(ce.get("openInterest", 0) or 0),
            "CE_ChgOI": int(ce.get("changeinOpenInterest", 0) or 0),
            "CE_LTP": float(ce.get("lastPrice", float("nan"))),
            "PE_OI": int(pe.get("openInterest", 0) or 0),
            "PE_ChgOI": int(pe.get("changeinOpenInterest", 0) or 0),
            "PE_LTP": float(pe.get("lastPrice", float("nan"))),
        })
    df = pd.DataFrame(rows).sort_values("strikePrice").reset_index(drop=True)
    ts_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return df, spot, expiry, ts_str

def calculate_pcr(df: pd.DataFrame) -> float:
    total_put_oi = float(df['PE_OI'].sum())
    total_call_oi = float(df['CE_OI'].sum())
    if total_call_oi == 0:
        return float('inf')
    return round(total_put_oi / total_call_oi, 2)

def calculate_max_pain(df: pd.DataFrame) -> int:
    strikes = np.sort(df['strikePrice'].dropna().unique())
    total_loss = {}
    for k in strikes:
        call_loss = (df.loc[df['strikePrice'] >= k, 'CE_OI'] * (df.loc[df['strikePrice'] >= k, 'strikePrice'] - k)).sum()
        put_loss = (df.loc[df['strikePrice'] <= k, 'PE_OI'] * (k - df.loc[df['strikePrice'] <= k, 'strikePrice'])).sum()
        total_loss[int(k)] = float(call_loss + put_loss)
    return int(min(total_loss, key=total_loss.get)) if total_loss else 0

def sr_levels_from_oi(df: pd.DataFrame) -> Tuple[int, int]:
    support = int(df.loc[df['PE_OI'].idxmax(), 'strikePrice']) if not df.empty else 0
    resistance = int(df.loc[df['CE_OI'].idxmax(), 'strikePrice']) if not df.empty else 0
    return support, resistance

def atm_straddle(df: pd.DataFrame, spot: float) -> Tuple[int, float, float, float]:
    if df.empty:
        return 0, float('nan'), float('nan'), float('nan')
    idx = (df['strikePrice'] - spot).abs().argsort().iloc[0]
    atm_k = int(df.iloc[idx]['strikePrice'])
    ce = float(df.iloc[idx].get('CE_LTP', float('nan')))
    pe = float(df.iloc[idx].get('PE_LTP', float('nan')))
    total = ce + pe if (not math.isnan(ce) and not math.isnan(pe)) else float('nan')
    return atm_k, ce, pe, total

def gamma_status(spot: float, support: int, resistance: int) -> str:
    if resistance and spot >= resistance + GAMMA_BREACH:
        return "Gamma Squeeze (Bullish)"
    if support and spot <= support - GAMMA_BREACH:
        return "Gamma Crash (Bearish)"
    return "Neutral"

# Max Pain strength classification AND numeric score
def max_pain_strength(df: pd.DataFrame, spot: float, max_pain: int) -> Tuple[str, float, int]:
    if df.empty or max_pain == 0:
        return "weak", 0.0, 0
    total_oi = float(df['CE_OI'].sum() + df['PE_OI'].sum()) or 1.0
    band = df[(df['strikePrice'] >= max_pain - ZONE_WIDTH) & (df['strikePrice'] <= max_pain + ZONE_WIDTH)]
    band_oi = float(band['CE_OI'].sum() + band['PE_OI'].sum())
    pct = (band_oi / total_oi) * 100.0

    dist = abs(spot - max_pain)
    if dist > FAR_FROM_MAXPAIN:
        return "breakout", pct, -3
    if pct >= 25.0:
        return "strong", pct, +2
    if pct >= 10.0:
        return "moderate", pct, +1
    return "weak", pct, 0

# ===================== SNAPSHOTS & ΔOI =====================
def _snapshot_fname(ts: str) -> str:
    return f"{SYMBOL}__{ts.replace(':', '').replace('-', '').replace(' ', '_')}.csv"

def save_snapshot(df: pd.DataFrame, ts: str) -> str:
    path = pathlib.Path(SNAPSHOT_DIR) / _snapshot_fname(ts)
    df.to_csv(path, index=False)
    return str(path)

def load_latest_snapshot() -> Optional[pd.DataFrame]:
    files = sorted(pathlib.Path(SNAPSHOT_DIR).glob("*.csv"))
    if not files:
        return None
    try:
        return pd.read_csv(files[-1])
    except Exception:
        return None

def load_snapshot_older_than(minutes: int) -> Tuple[Optional[pd.DataFrame], Optional[datetime]]:
    files = sorted(pathlib.Path(SNAPSHOT_DIR).glob("*.csv"))
    if not files:
        return None, None
    cutoff = datetime.now() - timedelta(minutes=minutes)
    chosen = None
    chosen_ts = None
    for f in reversed(files):
        try:
            ts_str = f.stem.split("__")[-1]
            ts = datetime.strptime(ts_str, "%Y%m%d_%H%M%S")
        except Exception:
            continue
        if ts <= cutoff:
            chosen, chosen_ts = f, ts
            break
    if chosen is None:
        return None, None
    try:
        return pd.read_csv(chosen), chosen_ts
    except Exception:
        return None, None

def compute_delta_oi(curr: pd.DataFrame, prev: Optional[pd.DataFrame]) -> pd.DataFrame:
    curr_slim = curr[['strikePrice', 'CE_OI', 'PE_OI']].copy()
    if prev is None or prev.empty:
        curr_slim['CE_DeltaOI'] = 0
        curr_slim['PE_DeltaOI'] = 0
        return curr_slim
    prev_slim = prev[['strikePrice', 'CE_OI', 'PE_OI']].copy()
    merged = pd.merge(curr_slim, prev_slim, on='strikePrice', how='left', suffixes=('', '_prev'))
    merged['CE_DeltaOI'] = (merged['CE_OI'] - merged['CE_OI_prev']).fillna(0).astype(int)
    merged['PE_DeltaOI'] = (merged['PE_OI'] - merged['PE_OI_prev']).fillna(0).astype(int)
    return merged[['strikePrice', 'CE_OI', 'PE_OI', 'CE_DeltaOI', 'PE_DeltaOI']]

def top_delta_oi(delta_df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    top_ce = delta_df.sort_values('CE_DeltaOI', ascending=False).head(TOP_N_DELTA_OI)
    top_pe = delta_df.sort_values('PE_DeltaOI', ascending=False).head(TOP_N_DELTA_OI)
    return top_ce.reset_index(drop=True), top_pe.reset_index(drop=True)

# ===================== BIAS (ATM ±3) =====================
def bias_table_atm_band(df: pd.DataFrame, spot: float, max_pain: int) -> Tuple[pd.DataFrame, int]:
    if df.empty:
        return pd.DataFrame(), 0
    # Find ATM strike
    atm_idx = (df['strikePrice'] - spot).abs().argsort().iloc[0]
    atm_strike = int(df.iloc[atm_idx]['strikePrice'])

    # Build strikes list: ATM ±3
    unique_strikes = df['strikePrice'].astype(int).unique()
    unique_strikes.sort()
    if atm_strike not in unique_strikes:
        # fallback just in case of float mismatch
        atm_strike = int(round(spot / 100.0) * 100)  # typical index step
    strikes_band = [atm_strike + i * 100 for i in range(-3, 4)]
    # Intersect with available strikes
    strikes_band = [k for k in strikes_band if k in set(unique_strikes)]
    
    rows = []
    total_score = 0
    for k in strikes_band:
        row = df.loc[df['strikePrice'] == k].iloc[0]
        ce_oi, pe_oi = int(row['CE_OI']), int(row['PE_OI'])
        ce_doi = int(row.get('CE_DeltaOI', 0)) if 'CE_DeltaOI' in row else 0
        pe_doi = int(row.get('PE_DeltaOI', 0)) if 'PE_DeltaOI' in row else 0
        
        # Base bias: more PE_OI => Bullish (+2), more CE_OI => Bearish (-2)
        if pe_oi > ce_oi:
            bias = 'Bullish'
            score = 2
        elif ce_oi > pe_oi:
            bias = 'Bearish'
            score = -2
        else:
            bias = 'Neutral'
            score = 0
        
        # ΔOI tilt: rising PE adds +1, rising CE adds -1 (use relative threshold)
        thr = max(100, int(0.02 * max(pe_oi + ce_oi, 1)))
        if pe_doi - ce_doi > thr:
            score += 1
        elif ce_doi - pe_doi > thr:
            score -= 1
        
        # Max Pain proximity (encourages reversion): if k == max_pain, add small neutral pull
        if k == max_pain:
            score += 1 if bias == 'Bullish' else (-1 if bias == 'Bearish' else 0)
        
        # Double-weight ATM
        if k == atm_strike:
            score *= 2
        
        rows.append({
            'Strike': k,
            'CE_OI': ce_oi,
            'CE_ΔOI': ce_doi,
            'PE_OI': pe_oi,
            'PE_ΔOI': pe_doi,
            'Bias': bias,
            'BiasScore': int(score),
        })
        total_score += int(score)
    
    bias_df = pd.DataFrame(rows)
    return bias_df, int(total_score)

# ===================== SIGNALS & SCORING =====================
def pcr_weighted_sr(support: int, resistance: int, pcr: float) -> Tuple[int, int, str]:
    bias = "Neutral"
    if pcr >= PCR_BULL:
        bias = "Support Bias (Bullish)"
    elif pcr <= PCR_BEAR:
        bias = "Resistance Bias (Bearish)"
    return support, resistance, bias

def zone_membership(spot: float, level: int) -> bool:
    return level > 0 and abs(spot - level) <= ZONE_WIDTH

def make_trade_signal(spot: float, pcr: float, max_pain_shift: Optional[int],
                      gamma: str, support: int, resistance: int,
                      bias_total: int) -> Dict[str, str]:
    """Generate signal only when alignments are met."""
    sig = {"Signal": "WAIT", "Side": "-", "EntryZone": "-", "SL": "-", "Target": "-", "Reason": "No alignment"}

    # CE candidate
    if (bias_total >= 2 and pcr > 1.0 and 
        (max_pain_shift is None or max_pain_shift >= 0) and 
        not gamma.startswith("Gamma Crash") and 
        zone_membership(spot, support)):
        entry = f"{support-ZONE_WIDTH} to {support+ZONE_WIDTH}"
        sl = round(support - 1.2 * ZONE_WIDTH, 1)
        tgt = round(support + 1.8 * ZONE_WIDTH, 1)
        return {
            "Signal": "Buy CE (Support Bounce)", 
            "Side": "CE", 
            "EntryZone": entry, 
            "SL": str(sl), 
            "Target": str(tgt), 
            "Reason": f"bias_total={bias_total}, PCR={pcr}>1, MP_shift≥0, {gamma}, in Support zone"
        }
    
    # PE candidate
    if (bias_total <= -2 and pcr < 1.0 and 
        (max_pain_shift is None or max_pain_shift <= 0) and 
        not gamma.startswith("Gamma Squeeze") and 
        zone_membership(spot, resistance)):
        entry = f"{resistance-ZONE_WIDTH} to {resistance+ZONE_WIDTH}"
        sl = round(resistance + 1.2 * ZONE_WIDTH, 1)
        tgt = round(resistance - 1.8 * ZONE_WIDTH, 1)
        return {
            "Signal": "Buy PE (Resistance Reject)", 
            "Side": "PE", 
            "EntryZone": entry, 
            "SL": str(sl), 
            "Target": str(tgt), 
            "Reason": f"bias_total={bias_total}, PCR={pcr}<1, MP_shift≤0, {gamma}, in Resistance zone"
        }
    
    return sig

def overall_score(pcr: float, mp_strength_score: int, gamma: str, bias_total: int, spot: float, max_pain: int) -> int:
    score = 50
    # PCR tilt
    if pcr >= PCR_BULL:
        score += 8
    elif pcr <= PCR_BEAR:
        score -= 8

    # Max pain strength numeric
    score += mp_strength_score  # (-3,0,+1,+2)
    
    # Gamma instability
    if gamma.startswith('Gamma Squeeze') or gamma.startswith('Gamma Crash'):
        score -= 5
    
    # Bias total add directly but clamp small range
    score += max(-10, min(10, bias_total))
    
    # Distance to Max Pain
    dist = abs(spot - max_pain)
    if dist <= 2 * ZONE_WIDTH:
        score += 4
    elif dist > FAR_FROM_MAXPAIN:
        score -= 4
    
    return max(0, min(100, int(score)))

# ===================== EXCEL I/O =====================
def append_to_excel(overview: pd.DataFrame,
                    chain_aug: pd.DataFrame,
                    top_ce: pd.DataFrame,
                    top_pe: pd.DataFrame,
                    bias_df: pd.DataFrame,
                    signal_row: pd.DataFrame):
    try:
        from openpyxl import load_workbook
    except ImportError:
        st.warning("openpyxl not installed. Excel functionality disabled.")
        return

    if not os.path.exists(EXCEL_FILE):
        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as w:
            overview.to_excel(w, sheet_name="Overview", index=False)
            chain_aug.to_excel(w, sheet_name="Chain", index=False)
            pd.concat([
                top_ce.assign(Type='CEΔOI'),
                top_pe.assign(Type='PEΔOI')
            ], ignore_index=True).to_excel(w, sheet_name="TopDeltaOI", index=False)
            bias_df.to_excel(w, sheet_name="BiasScoring", index=False)
            signal_row.to_excel(w, sheet_name="SignalsLog", index=False)
        return

    book = load_workbook(EXCEL_FILE)
    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="overlay") as w:
        # Overview append
        if "Overview" in book.sheetnames:
            start_row = book["Overview"].max_row
            overview.to_excel(w, sheet_name="Overview", index=False, header=False, startrow=start_row)
        else:
            overview.to_excel(w, sheet_name="Overview", index=False)
        
        # Chain replace with latest snapshot
        if "Chain" in book.sheetnames:
            del book["Chain"]
            book.create_sheet("Chain")
        chain_aug.to_excel(w, sheet_name="Chain", index=False)
        
        # Top ΔOI replace
        if "TopDeltaOI" in book.sheetnames:
            del book["TopDeltaOI"]
            book.create_sheet("TopDeltaOI")
        pd.concat([
            top_ce.assign(Type='CEΔOI'),
            top_pe.assign(Type='PEΔOI')
        ], ignore_index=True).to_excel(w, sheet_name="TopDeltaOI", index=False)
        
        # BiasScoring replace
        if "BiasScoring" in book.sheetnames:
            del book["BiasScoring"]
            book.create_sheet("BiasScoring")
        bias_df.to_excel(w, sheet_name="BiasScoring", index=False)
        
        # SignalsLog append
        if "SignalsLog" in book.sheetnames:
            start_row = book["SignalsLog"].max_row
            signal_row.to_excel(w, sheet_name="SignalsLog", index=False, header=False, startrow=start_row)
        else:
            signal_row.to_excel(w, sheet_name="SignalsLog", index=False)

# ===================== MAIN ANALYSIS FUNCTION =====================
def run_analysis():
    try:
        if not is_valid_trading_time():
            st.warning("Outside trading hours (Mon-Fri, 9:00-15:30). Analysis paused.")
            return None

        raw = fetch_option_chain()
        df, spot, expiry, ts = parse_chain(raw)

        # core metrics
        pcr = calculate_pcr(df)
        max_pain = calculate_max_pain(df)
        support, resistance = sr_levels_from_oi(df)
        support, resistance, sr_bias = pcr_weighted_sr(support, resistance, pcr)
        atm_k, atm_ce, atm_pe, atm_total = atm_straddle(df, spot)
        gamma = gamma_status(spot, support, resistance)
        mp_strength_label, mp_strength_pct, mp_strength_score = max_pain_strength(df, spot, max_pain)
        
        # ΔOI vs latest snapshot
        prev_df = load_latest_snapshot()
        delta_df = compute_delta_oi(df, prev_df)
        top_ce, top_pe = top_delta_oi(delta_df)
        
        # merge ΔOI into chain for excel
        chain_aug = pd.merge(df, delta_df, on='strikePrice', how='left')
        chain_aug.insert(0, 'Timestamp', ts)
        
        # max pain shift vs ~1 hour ago
        hour_df, hour_ts = load_snapshot_older_than(55)
        if hour_df is not None:
            mp_hour = calculate_max_pain(hour_df)
            mp_shift = int(max_pain - mp_hour)
        else:
            mp_shift = None
        
        # Bias table (ATM ±3) using chain_aug (has delta columns)
        bias_df, bias_total = bias_table_atm_band(chain_aug, spot, max_pain)
        bias_df.insert(0, 'Timestamp', ts)
        
        # overall score and signal
        total_score = overall_score(pcr, mp_strength_score, gamma, bias_total, spot, max_pain)
        signal = make_trade_signal(spot, pcr, mp_shift, gamma, support, resistance, bias_total)
        
        # Overview row
        overview = pd.DataFrame([{
            "Timestamp": ts,
            "Symbol": SYMBOL,
            "Expiry": expiry,
            "Spot": spot,
            "PCR": pcr,
            "MaxPain": max_pain,
            "MaxPainShift_vs_1h": mp_shift if mp_shift is not None else "NA",
            "MaxPainStrength": mp_strength_label,
            "MaxPainStrength%": round(mp_strength_pct, 2),
            "MaxPainStrengthScore": mp_strength_score,
            "Support": support,
            "Resistance": resistance,
            "SR_Bias": sr_bias,
            "ZoneWidth": ZONE_WIDTH,
            "Gamma": gamma,
            "ATM_Strike": atm_k,
            "ATM_CE_LTP": atm_ce,
            "ATM_PE_LTP": atm_pe,
            "ATM_Straddle": atm_total,
            "BiasTotal(ATM±3)": bias_total,
            "Score": total_score,
            "Signal": signal.get("Signal"),
            "Side": signal.get("Side"),
            "EntryZone": signal.get("EntryZone"),
            "SL": signal.get("SL"),
            "Target": signal.get("Target"),
            "Reason": signal.get("Reason"),
        }])
        
        # Signals log row (for append)
        signals_log = overview[[
            "Timestamp", "Symbol", "Spot", "PCR", "MaxPain", "MaxPainShift_vs_1h",
            "MaxPainStrength", "MaxPainStrength%", "MaxPainStrengthScore",
            "Support", "Resistance", "Gamma", "ATM_Strike", "ATM_Straddle",
            "BiasTotal(ATM±3)", "Score", "Signal", "Side", "EntryZone", "SL", "Target", "Reason"
        ]].copy()
        
        # Save snapshot and excel
        snap_path = save_snapshot(df, ts)
        append_to_excel(overview, chain_aug, top_ce, top_pe, bias_df, signals_log)
        
        # Send Telegram notification if there's a signal
        if signal["Signal"] != "WAIT":
            telegram_msg = f"{SYMBOL} Signal: {signal['Signal']}\nSpot: {spot}, PCR: {pcr}\nEntry: {signal['EntryZone']}\nReason: {signal['Reason']}"
            send_telegram_message(telegram_msg)
        
        return {
            "overview": overview,
            "chain_aug": chain_aug,
            "top_ce": top_ce,
            "top_pe": top_pe,
            "bias_df": bias_df,
            "signals_log": signals_log,
            "snap_path": snap_path,
            "timestamp": ts
        }
        
    except requests.exceptions.HTTPError as he:
        st.error(f"HTTP error: {he}")
    except requests.exceptions.RequestException as re:
        st.error(f"Network error: {re}")
    except Exception as e:
        st.error(f"Error: {e}")
    
    return None

# ===================== STREAMLIT UI =====================
def main():
    st.set_page_config(
        page_title=f"{SYMBOL} Expiry Day Analytics",
        page_icon="📈",
        layout="wide"
    )
    
    st.title(f"{SYMBOL} Expiry Day Analytics Dashboard")
    st.write("Auto-refreshing every 2 minutes | Mon-Fri 9:00-15:30")
    
    # Initialize session state
    if 'last_update' not in st.session_state:
        st.session_state.last_update = None
    if 'data' not in st.session_state:
        st.session_state.data = None
    
    # Auto-refresh logic
    refresh_interval = 120  # 2 minutes
    if st.session_state.last_update is None or (datetime.now() - st.session_state.last_update).seconds >= refresh_interval:
        with st.spinner("Fetching latest data..."):
            st.session_state.data = run_analysis()
            st.session_state.last_update = datetime.now()
    
    # Display data if available
    if st.session_state.data:
        data = st.session_state.data
        overview = data["overview"]
        
        # Display key metrics
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Spot", overview["Spot"].iloc[0])
            st.metric("PCR", overview["PCR"].iloc[0])
        with col2:
            st.metric("Max Pain", overview["MaxPain"].iloc[0])
            st.metric("Max Pain Strength", overview["MaxPainStrength"].iloc[0])
        with col3:
            st.metric("Support", overview["Support"].iloc[0])
            st.metric("Resistance", overview["Resistance"].iloc[0])
        with col4:
            st.metric("Signal", overview["Signal"].iloc[0])
            st.metric("Overall Score", overview["Score"].iloc[0])
        
        # Display detailed sections in tabs
        tab1, tab2, tab3, tab4 = st.tabs(["Overview", "Bias Analysis", "Top ΔOI", "Chain Data"])
        
        with tab1:
            st.subheader("Overview")
            st.dataframe(overview)
            
            st.subheader("Signal Details")
            if overview["Signal"].iloc[0] != "WAIT":
                st.success(f"**{overview['Signal'].iloc[0]}**")
                st.write(f"**Entry Zone:** {overview['EntryZone'].iloc[0]}")
                st.write(f"**Stop Loss:** {overview['SL'].iloc[0]}")
                st.write(f"**Target:** {overview['Target'].iloc[0]}")
                st.write(f"**Reason:** {overview['Reason'].iloc[0]}")
            else:
                st.info("No strong signal detected. Waiting for better alignment.")
        
        with tab2:
            st.subheader("Bias Analysis (ATM ±3 Strikes)")
            st.dataframe(data["bias_df"])
            st.metric("Total Bias Score", overview["BiasTotal(ATM±3)"].iloc[0])
        
        with tab3:
            st.subheader("Top Call ΔOI")
            st.dataframe(data["top_ce"])
            st.subheader("Top Put ΔOI")
            st.dataframe(data["top_pe"])
        
        with tab4:
            st.subheader("Option Chain Data")
            st.dataframe(data["chain_aug"].head(20))
        
        st.write(f"Last updated: {data['timestamp']}")
        st.write(f"Snapshot saved: {data['snap_path']}")
    
    # Manual refresh button
    if st.button("Refresh Now"):
        with st.spinner("Refreshing data..."):
            st.session_state.data = run_analysis()
            st.session_state.last_update = datetime.now()
        st.rerun()
    
    # Countdown to next refresh
    if st.session_state.last_update:
        next_refresh = st.session_state.last_update + timedelta(seconds=refresh_interval)
        time_remaining = next_refresh - datetime.now()
        st.write(f"Next refresh in: {time_remaining.seconds} seconds")

if __name__ == "__main__":
    main()
